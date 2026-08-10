import csv
import io
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from ..database import get_db
from .. import models, security, attendance_service

router = APIRouter(prefix="/api/export", tags=["export"])


def _get_records(db: Session, record_date, student_id, class_id, subject_id=None):
    q = db.query(models.AttendanceRecord)
    if record_date:
        q = q.filter(models.AttendanceRecord.date == record_date)
    if student_id:
        q = q.filter(models.AttendanceRecord.student_id == student_id)
    if class_id:
        q = q.filter(models.AttendanceRecord.session_id.in_(
            db.query(models.AttendanceSession.id).filter(models.AttendanceSession.class_id == class_id)))
    if subject_id:
        q = q.filter(models.AttendanceRecord.session_id.in_(
            db.query(models.AttendanceSession.id).filter(models.AttendanceSession.subject_id == subject_id)))
    return q.all()


def _rows(records):
    return [["ID", "Student", "Subject", "Class", "Teacher", "Status", "Date", "Time", "Confidence", "Method"]] + [
        [r.id, r.student_name, r.subject, r.class_name, r.teacher, r.status, str(r.date), r.time, r.confidence, r.method]
        for r in records
    ]


def _get_student(db: Session, user: models.User) -> models.Student:
    """Get the student record for the current user."""
    student = db.query(models.Student).filter(models.Student.user_id == user.id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student profile not found")
    return student


def _student_report_data(db: Session, student: models.Student):
    """Build student attendance report data."""
    records = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.student_id == student.id
    ).order_by(models.AttendanceRecord.date.desc(), models.AttendanceRecord.time.desc()).all()

    total = len(records)
    present = len([r for r in records if r.status == "present"])
    absent = len([r for r in records if r.status == "absent"])
    late = len([r for r in records if r.status == "late"])
    percentage = round(present / total * 100, 2) if total else 0.0

    # Subject-wise breakdown
    from collections import defaultdict
    subj_map = defaultdict(lambda: {"total": 0, "present": 0, "absent": 0, "late": 0})
    for r in records:
        key = r.subject or "General"
        subj_map[key]["total"] += 1
        if r.status == "present":
            subj_map[key]["present"] += 1
        elif r.status == "absent":
            subj_map[key]["absent"] += 1
        elif r.status == "late":
            subj_map[key]["late"] += 1
    subjects = [{"name": k, **v,
                 "percentage": round(v["present"] / v["total"] * 100, 2) if v["total"] else 0}
                for k, v in subj_map.items()]

    return {
        "student": student,
        "records": records,
        "total": total,
        "present": present,
        "absent": absent,
        "late": late,
        "percentage": percentage,
        "subjects": subjects,
    }


@router.get("/csv")
def export_csv(record_date: date = None, student_id: int = None, class_id: int = None,
               subject_id: int = None, user: models.User = Depends(security.get_current_user),
               db: Session = Depends(get_db)):
    records = _get_records(db, record_date, student_id, class_id, subject_id)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(_rows(records))
    buf.seek(0)
    filename = f"attendance_{date.today()}.csv"
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/excel")
def export_excel(record_date: date = None, student_id: int = None, class_id: int = None,
                 subject_id: int = None, user: models.User = Depends(security.get_current_user),
                 db: Session = Depends(get_db)):
    from openpyxl import Workbook
    records = _get_records(db, record_date, student_id, class_id, subject_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    for row in _rows(records):
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"attendance_{date.today()}.xlsx"
    return StreamingResponse(iter([buf.getvalue()]),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/pdf")
def export_pdf(record_date: date = None, student_id: int = None, class_id: int = None,
               subject_id: int = None, user: models.User = Depends(security.get_current_user),
               db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    records = _get_records(db, record_date, student_id, class_id, subject_id)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = [Paragraph("Attendance Report - LifeOS Smart Campus", styles["Title"]), Spacer(1, 12)]
    data = _rows(records)
    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366f1")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    filename = f"attendance_{date.today()}.pdf"
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


# ============ STUDENT-SPECIFIC REPORT DOWNLOADS ============

@router.get("/student-report")
def student_report(format: str = "pdf",
                   user: models.User = Depends(security.require_roles("student")),
                   db: Session = Depends(get_db)):
    """Download the logged-in student's attendance report (PDF/Excel/CSV)."""
    student = _get_student(db, user)
    data = _student_report_data(db, student)

    if format == "csv":
        return _student_csv(data)
    elif format == "xlsx" or format == "excel":
        return _student_excel(data)
    elif format == "pdf":
        return _student_pdf(data)
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported format: {format}. Use pdf, xlsx, or csv.")


def _student_csv(data):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["LifeOS Smart Campus - Student Attendance Report"])
    writer.writerow([])
    writer.writerow(["Student Name", data["student"].full_name])
    writer.writerow(["Roll Number", data["student"].roll_number])
    writer.writerow(["Student ID", data["student"].student_id])
    writer.writerow(["Department", data["student"].department.name if data["student"].department else ""])
    writer.writerow(["Course", data["student"].course.name if data["student"].course else ""])
    writer.writerow(["Semester", data["student"].semester.name if data["student"].semester else ""])
    writer.writerow(["Class", data["student"].class_.name if data["student"].class_ else ""])
    writer.writerow(["Attendance Percentage", f"{data['percentage']}%"])
    writer.writerow([])
    writer.writerow(["Subject", "Total", "Present", "Absent", "Late", "Percentage"])
    for s in data["subjects"]:
        writer.writerow([s["name"], s["total"], s["present"], s["absent"], s["late"], f"{s['percentage']}%"])
    writer.writerow([])
    writer.writerow(["Date", "Time", "Subject", "Teacher", "Status", "Confidence"])
    for r in data["records"]:
        writer.writerow([str(r.date), r.time, r.subject, r.teacher, r.status, r.confidence])
    buf.seek(0)
    filename = f"student_attendance_{data['student'].student_id}_{date.today()}.csv"
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


def _student_excel(data):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = "LifeOS Smart Campus - Student Attendance Report"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center")

    # Student info
    ws["A3"] = "Student Name"
    ws["B3"] = data["student"].full_name
    ws["A4"] = "Roll Number"
    ws["B4"] = data["student"].roll_number
    ws["A5"] = "Student ID"
    ws["B5"] = data["student"].student_id
    ws["A6"] = "Department"
    ws["B6"] = data["student"].department.name if data["student"].department else ""
    ws["A7"] = "Course"
    ws["B7"] = data["student"].course.name if data["student"].course else ""
    ws["A8"] = "Semester"
    ws["B8"] = data["student"].semester.name if data["student"].semester else ""
    ws["A9"] = "Class"
    ws["B9"] = data["student"].class_.name if data["student"].class_ else ""
    ws["A10"] = "Attendance Percentage"
    ws["B10"] = f"{data['percentage']}%"

    # Subject-wise
    ws["A12"] = "Subject-wise Attendance"
    ws["A12"].font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["Subject", "Total", "Present", "Absent", "Late", "Percentage"])
    for s in data["subjects"]:
        ws.append([s["name"], s["total"], s["present"], s["absent"], s["late"], f"{s['percentage']}%"])

    # History
    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="Attendance History").font = Font(bold=True, size=12)
    row += 1
    headers = ["Date", "Time", "Subject", "Teacher", "Status", "Confidence"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h).font = Font(bold=True)
    for r in data["records"]:
        row += 1
        ws.cell(row=row, column=1, value=str(r.date))
        ws.cell(row=row, column=2, value=r.time)
        ws.cell(row=row, column=3, value=r.subject)
        ws.cell(row=row, column=4, value=r.teacher)
        ws.cell(row=row, column=5, value=r.status)
        ws.cell(row=row, column=6, value=r.confidence)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"student_attendance_{data['student'].student_id}_{date.today()}.xlsx"
    return StreamingResponse(iter([buf.getvalue()]),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


def _student_pdf(data):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#6366f1"),
        alignment=1, spaceAfter=6
    )
    elements = [Paragraph("LifeOS Smart Campus", title_style),
                Paragraph("Student Attendance Report", styles["Heading2"]),
                Spacer(1, 12)]

    # Student info table
    s = data["student"]
    info_data = [
        ["Student Name", s.full_name, "Roll Number", s.roll_number],
        ["Student ID", s.student_id, "Department", s.department.name if s.department else ""],
        ["Course", s.course.name if s.course else "", "Semester", s.semester.name if s.semester else ""],
        ["Class", s.class_.name if s.class_ else "", "Attendance %", f"{data['percentage']}%"],
    ]
    info_table = Table(info_data, colWidths=[1.2 * inch, 2.3 * inch, 1.2 * inch, 2.3 * inch])
    info_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f3f4f6")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#f3f4f6")),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 16))

    # Subject-wise
    elements.append(Paragraph("Subject-wise Attendance", styles["Heading3"]))
    subj_data = [["Subject", "Total", "Present", "Absent", "Late", "Percentage"]]
    for subj in data["subjects"]:
        subj_data.append([subj["name"], subj["total"], subj["present"], subj["absent"], subj["late"], f"{subj['percentage']}%"])
    subj_table = Table(subj_data)
    subj_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    elements.append(subj_table)
    elements.append(Spacer(1, 16))

    # History
    elements.append(Paragraph("Attendance History", styles["Heading3"]))
    hist_data = [["Date", "Time", "Subject", "Teacher", "Status", "Confidence"]]
    for r in data["records"]:
        hist_data.append([str(r.date), r.time, r.subject, r.teacher, r.status, r.confidence])
    hist_table = Table(hist_data)
    hist_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366f1")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f3f4f6")]),
    ]))
    elements.append(hist_table)

    doc.build(elements)
    buf.seek(0)
    filename = f"student_attendance_{data['student'].student_id}_{date.today()}.pdf"
    return StreamingResponse(iter([buf.getvalue()]), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})